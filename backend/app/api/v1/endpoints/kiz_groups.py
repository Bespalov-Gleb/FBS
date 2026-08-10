from __future__ import annotations

import csv
import io
from collections import defaultdict
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import CurrentAdminUser
from app.models.kiz_group import KizGroup, kiz_group_marketplaces
from app.models.kiz_parser_error import KizParserError
from app.models.kiz_pool_item import KizCodeStatus, KizPoolItem
from app.models.kiz_product_mapping import KizProductMapping
from app.models.kiz_settings import KizSettings
from app.models.marketplace import Marketplace, MarketplaceType
from app.models.order import Order
from app.models.print_settings import PrintSettings
from app.models.user import User
from app.services.kiz_pool_service import (
    import_kiz_codes_from_pdfs,
    release_kiz_pool_item_to_free,
    take_free_kiz_codes_for_manual_print,
)

router = APIRouter(prefix="/kiz-groups", tags=["KIZ Groups"])


class KizGroupRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    color: Optional[str] = Field(default=None, max_length=120)
    size: Optional[str] = Field(default=None, max_length=120)
    cut_type: Optional[str] = Field(default=None, max_length=120)
    parser_markers: Optional[dict[str, Any]] = None
    marketplace_ids: list[int] = Field(default_factory=list)


class KizGroupResponse(BaseModel):
    id: int
    name: str
    color: Optional[str]
    size: Optional[str]
    cut_type: Optional[str]
    parser_markers: Optional[dict[str, Any]]
    marketplace_ids: list[int]
    free_count: int
    used_count: int
    parser_errors_count: int


class ProductGroupMappingUpsert(BaseModel):
    marketplace_id: int
    article: str
    size: Optional[str] = ""
    group_id: int


class PrintFromGroupRequest(BaseModel):
    count: int = Field(..., ge=1, description="Сколько КИЗ списать и распечатать")


class ColorMarkersRequest(BaseModel):
    color_markers: list[str] = Field(default_factory=list)


class ColorMarkersResponse(BaseModel):
    color_markers: list[str]


DEFAULT_COLOR_MARKERS = [
    "manblack",
    "manwhite",
    "girlblack",
    "girlwhite",
    "womanblack",
    "womanwhite",
    "black",
    "white",
    "beige",
    "gray",
    "grey",
    "blue",
    "red",
    "green",
    "pink",
    "brown",
    "yellow",
    "orange",
    "purple",
    "navy",
    "cream",
    "khaki",
]


def _normalize_color_markers(raw: list[str] | None) -> list[str]:
    cleaned: list[str] = []
    seen: set[str] = set()
    for item in raw or []:
        marker = str(item or "").strip()
        if not marker:
            continue
        key = marker.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(marker)
    return cleaned


def _get_or_create_kiz_settings(db: Session, user_id: int) -> KizSettings:
    row = db.query(KizSettings).filter(KizSettings.user_id == user_id).first()
    if row:
        return row
    row = KizSettings(user_id=user_id, color_markers=list(DEFAULT_COLOR_MARKERS))
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _split_article_and_size(article: str | None) -> tuple[str, str]:
    """
    Хвост артикула по последнему '_' (например ABC123_XL).
    Если шаблон не распознан, считаем, что хвоста нет.
    """
    raw = (article or "").strip()
    if not raw:
        return "", ""
    if "_" not in raw:
        return raw, ""
    base, size = raw.rsplit("_", 1)
    base = base.strip()
    size = size.strip()
    if not base or not size:
        return raw, ""
    return base, size


def _is_wildberries(marketplace_type: MarketplaceType | str | None) -> bool:
    if marketplace_type is None:
        return False
    if isinstance(marketplace_type, MarketplaceType):
        return marketplace_type == MarketplaceType.WILDBERRIES
    return str(marketplace_type).lower() == MarketplaceType.WILDBERRIES.value


_GENDER_COLOR_PREFIXES = ("woman", "girl", "unisex", "male", "female", "man")


def _normalize_color_token(raw: str | None) -> str:
    """
    Убрать пол из хвоста артикула WB: manblack → black, girlwhite → white.
    Если после среза ничего не осталось — вернуть исходное значение.
    """
    s = (raw or "").strip()
    if not s:
        return ""
    lower = s.lower()
    for prefix in sorted(_GENDER_COLOR_PREFIXES, key=len, reverse=True):
        if lower.startswith(prefix) and len(s) > len(prefix):
            return s[len(prefix) :]
    return s


def _color_from_markers(article: str | None, markers: list[str] | None) -> str:
    """Первая (самая длинная) подстрока цвета, найденная в артикуле."""
    hay = (article or "").lower()
    if not hay:
        return ""
    ordered = sorted(
        _normalize_color_markers(markers),
        key=lambda m: len(m),
        reverse=True,
    )
    for marker in ordered:
        if marker.lower() in hay:
            return marker
    return ""


def _strip_color_from_article_base(article_base: str, color: str) -> str:
    """Если база оканчивается на _color — убрать хвост для более чистого артикула."""
    base = (article_base or "").strip()
    c = (color or "").strip()
    if not base or not c:
        return base
    suffix = f"_{c}"
    if base.lower().endswith(suffix.lower()) and len(base) > len(suffix):
        return base[: -len(suffix)].rstrip("_") or base
    return base


def _article_display_parts(
    article: str | None,
    *,
    marketplace_type: MarketplaceType | str | None,
    size_from_order: str | None,
    color_markers: list[str] | None = None,
) -> tuple[str, str, str]:
    """
    (article_base, size, color) для таблицы привязки товаров.

    WB: хвост артикула — цвет (без префикса пола), иначе поиск по color_markers;
         размер — из extra_data.size.
    Ozon: размер из order_size/хвоста; цвет — поиск color_markers в полном артикуле.
    """
    raw = (article or "").strip()
    article_base, suffix = _split_article_and_size(raw)
    order_size = (size_from_order or "").strip()
    marker_color = _color_from_markers(raw, color_markers)

    if _is_wildberries(marketplace_type):
        color = _normalize_color_token(suffix) or marker_color
        if color:
            article_base = _strip_color_from_article_base(article_base, color)
            # хвост мог быть manblack, а color уже black — тоже срезать исходный suffix
            if suffix and suffix.lower() != color.lower():
                article_base = _strip_color_from_article_base(article_base, suffix)
        return article_base, order_size, color

    size = order_size or suffix
    color = marker_color
    if color:
        # у Ozon цвет часто в середине/хвосте: cepi_cepi_white_L → base без white
        article_base = _strip_color_from_article_base(article_base, color)
    return article_base, size, color


def _find_product_mapping(
    db: Session,
    *,
    user_id: int,
    marketplace_id: int,
    article: str,
    size: str,
    color: str = "",
):
    """Ищем маппинг по реальному размеру; для WB — фолбэк на старый маппинг по цвету в size."""
    mapping = (
        db.query(KizProductMapping, KizGroup.name)
        .join(KizGroup, KizGroup.id == KizProductMapping.group_id)
        .filter(
            KizProductMapping.user_id == user_id,
            KizProductMapping.marketplace_id == marketplace_id,
            KizProductMapping.article == article,
            KizProductMapping.size == size,
        )
        .first()
    )
    if mapping or not color or color == size:
        return mapping
    return (
        db.query(KizProductMapping, KizGroup.name)
        .join(KizGroup, KizGroup.id == KizProductMapping.group_id)
        .filter(
            KizProductMapping.user_id == user_id,
            KizProductMapping.marketplace_id == marketplace_id,
            KizProductMapping.article == article,
            KizProductMapping.size == color,
        )
        .first()
    )


def _group_to_response(
    group: KizGroup,
    *,
    free_count: int,
    used_count: int,
    parser_errors_count: int,
) -> KizGroupResponse:
    return KizGroupResponse(
        id=group.id,
        name=group.name,
        color=group.color,
        size=group.size,
        cut_type=group.cut_type,
        parser_markers=group.parser_markers,
        marketplace_ids=[m.id for m in group.marketplaces],
        free_count=free_count,
        used_count=used_count,
        parser_errors_count=parser_errors_count,
    )


@router.get("", response_model=list[KizGroupResponse])
def list_kiz_groups(
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    groups = (
        db.query(KizGroup)
        .filter(KizGroup.user_id == current_user.id)
        .order_by(KizGroup.created_at.desc())
        .all()
    )
    if not groups:
        return []

    group_ids = [g.id for g in groups]
    pool_counts = (
        db.query(KizPoolItem.group_id, KizPoolItem.status, func.count(KizPoolItem.id))
        .filter(KizPoolItem.group_id.in_(group_ids))
        .group_by(KizPoolItem.group_id, KizPoolItem.status)
        .all()
    )
    errors_counts = (
        db.query(KizParserError.group_id, func.count(KizParserError.id))
        .filter(KizParserError.group_id.in_(group_ids))
        .group_by(KizParserError.group_id)
        .all()
    )

    free_by_group: dict[int, int] = defaultdict(int)
    used_by_group: dict[int, int] = defaultdict(int)
    err_by_group: dict[int, int] = defaultdict(int)

    for gid, status_val, cnt in pool_counts:
        normalized_status = status_val.value if hasattr(status_val, "value") else str(status_val)
        if normalized_status == KizCodeStatus.FREE.value:
            free_by_group[gid] = int(cnt)
        elif normalized_status == KizCodeStatus.USED.value:
            used_by_group[gid] = int(cnt)
    for gid, cnt in errors_counts:
        err_by_group[gid] = int(cnt)

    return [
        _group_to_response(
            g,
            free_count=free_by_group[g.id],
            used_count=used_by_group[g.id],
            parser_errors_count=err_by_group[g.id],
        )
        for g in groups
    ]


@router.get("/color-markers", response_model=ColorMarkersResponse)
def get_color_markers(
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    settings = _get_or_create_kiz_settings(db, current_user.id)
    markers = _normalize_color_markers(settings.color_markers)
    if not markers:
        markers = list(DEFAULT_COLOR_MARKERS)
    return ColorMarkersResponse(color_markers=markers)


@router.put("/color-markers", response_model=ColorMarkersResponse)
def update_color_markers(
    payload: ColorMarkersRequest,
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    settings = _get_or_create_kiz_settings(db, current_user.id)
    settings.color_markers = _normalize_color_markers(payload.color_markers)
    db.commit()
    db.refresh(settings)
    return ColorMarkersResponse(color_markers=_normalize_color_markers(settings.color_markers))


@router.post("", response_model=KizGroupResponse)
def create_kiz_group(
    payload: KizGroupRequest,
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Название группы обязательно.")

    exists = (
        db.query(KizGroup.id)
        .filter(KizGroup.user_id == current_user.id, KizGroup.name == name)
        .first()
    )
    if exists:
        raise HTTPException(status_code=400, detail="Группа с таким названием уже существует.")

    marketplaces = []
    if payload.marketplace_ids:
        marketplaces = (
            db.query(Marketplace)
            .filter(
                Marketplace.user_id == current_user.id,
                Marketplace.id.in_(payload.marketplace_ids),
            )
            .all()
        )
        if len(marketplaces) != len(set(payload.marketplace_ids)):
            raise HTTPException(status_code=400, detail="Часть магазинов не найдена.")

    group = KizGroup(
        user_id=current_user.id,
        name=name,
        color=(payload.color or "").strip() or None,
        size=(payload.size or "").strip() or None,
        cut_type=(payload.cut_type or "").strip() or None,
        parser_markers=payload.parser_markers,
        is_active=True,
    )
    group.marketplaces = marketplaces
    db.add(group)
    db.commit()
    db.refresh(group)
    return _group_to_response(group, free_count=0, used_count=0, parser_errors_count=0)


@router.patch("/{group_id}", response_model=KizGroupResponse)
def update_kiz_group(
    group_id: int,
    payload: KizGroupRequest,
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    group = (
        db.query(KizGroup)
        .filter(KizGroup.id == group_id, KizGroup.user_id == current_user.id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена.")

    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Название группы обязательно.")

    dup = (
        db.query(KizGroup.id)
        .filter(
            KizGroup.user_id == current_user.id,
            KizGroup.name == name,
            KizGroup.id != group_id,
        )
        .first()
    )
    if dup:
        raise HTTPException(status_code=400, detail="Группа с таким названием уже существует.")

    marketplaces = (
        db.query(Marketplace)
        .filter(
            Marketplace.user_id == current_user.id,
            Marketplace.id.in_(payload.marketplace_ids or []),
        )
        .all()
    )
    if len(marketplaces) != len(set(payload.marketplace_ids or [])):
        raise HTTPException(status_code=400, detail="Часть магазинов не найдена.")

    group.name = name
    group.color = (payload.color or "").strip() or None
    group.size = (payload.size or "").strip() or None
    group.cut_type = (payload.cut_type or "").strip() or None
    group.parser_markers = payload.parser_markers
    group.marketplaces = marketplaces
    db.commit()
    db.refresh(group)

    free_count = (
        db.query(func.count(KizPoolItem.id))
        .filter(KizPoolItem.group_id == group.id, KizPoolItem.status == KizCodeStatus.FREE)
        .scalar()
        or 0
    )
    used_count = (
        db.query(func.count(KizPoolItem.id))
        .filter(KizPoolItem.group_id == group.id, KizPoolItem.status == KizCodeStatus.USED)
        .scalar()
        or 0
    )
    parser_errors_count = (
        db.query(func.count(KizParserError.id))
        .filter(KizParserError.group_id == group.id)
        .scalar()
        or 0
    )

    return _group_to_response(
        group,
        free_count=int(free_count),
        used_count=int(used_count),
        parser_errors_count=int(parser_errors_count),
    )


@router.post("/{group_id}/upload-pdf")
async def upload_kiz_pdfs(
    group_id: int,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    group = (
        db.query(KizGroup)
        .filter(KizGroup.id == group_id, KizGroup.user_id == current_user.id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена.")
    if not files:
        raise HTTPException(status_code=400, detail="Не выбраны PDF файлы.")

    loaded_files: list[tuple[str, bytes]] = []
    for uploaded in files:
        filename = uploaded.filename or "file.pdf"
        if Path(filename).suffix.lower() != ".pdf":
            raise HTTPException(
                status_code=400,
                detail=f"Файл {filename}: поддерживаются только PDF.",
            )
        loaded_files.append((filename, await uploaded.read()))

    try:
        stats = import_kiz_codes_from_pdfs(
            db,
            owner_user_id=current_user.id,
            group=group,
            uploaded_files=loaded_files,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка импорта PDF: {exc}") from exc

    return {
        "ok": True,
        "total_pages": stats.total_pages,
        "imported": stats.imported,
        "duplicates": stats.duplicates,
        "errors": stats.errors,
    }


@router.post("/{group_id}/print")
def print_kiz_from_group(
    group_id: int,
    payload: PrintFromGroupRequest,
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    """
    Списать N свободных КИЗ из группы и вернуть multipage PDF этикеток для печати.
    Кривые коды пропускаются и возвращаются в остаток; пачка не валится целиком.
    """
    group = (
        db.query(KizGroup)
        .filter(KizGroup.id == group_id, KizGroup.user_id == current_user.id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена.")

    try:
        pool_items = take_free_kiz_codes_for_manual_print(
            db,
            group=group,
            count=payload.count,
            used_by_user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    # Lazy import после списания-валидации: генерация DataMatrix / merge PDF.
    from app.api.v1.endpoints.orders import (
        _generate_kiz_label_pdf,
        _kiz_31,
        _normalize_kiz_for_label,
        _rotate_pdf,
    )
    from pypdf import PdfReader, PdfWriter

    ps = db.query(PrintSettings).filter(PrintSettings.user_id == current_user.id).first()
    kiz_w = (ps.kiz_width_mm or 40) if ps else 40
    kiz_h = (ps.kiz_height_mm or 35) if ps else 35
    kiz_rot = (ps.kiz_rotate or 0) if ps else 0

    writer = PdfWriter()
    printed = 0
    skipped_errors: list[str] = []

    for item in pool_items:
        try:
            kiz = _normalize_kiz_for_label(item.code)
            kiz_31 = _kiz_31(kiz)
            if not kiz_31:
                raise ValueError("пустой КИЗ после нормализации")
            page_pdf = _generate_kiz_label_pdf(kiz, kiz_31, width_mm=kiz_w, height_mm=kiz_h)
            if kiz_rot:
                page_pdf = _rotate_pdf(page_pdf, kiz_rot)
            reader = PdfReader(io.BytesIO(page_pdf))
            for page in reader.pages:
                writer.add_page(page)
            printed += 1
        except Exception as exc:
            release_kiz_pool_item_to_free(item)
            preview = (item.code or "")[:24]
            skipped_errors.append(f"{preview}… ({exc})")

    if printed == 0:
        db.rollback()
        detail = "Не удалось сформировать ни одной этикетки КИЗ."
        if skipped_errors:
            detail += " Примеры: " + "; ".join(skipped_errors[:3])
        raise HTTPException(status_code=400, detail=detail)

    db.commit()

    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    safe_group = "".join(ch for ch in (group.name or "group")[:40] if ch.isalnum() or ch in "-_") or "group"
    skipped = len(skipped_errors)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="kiz-print-{safe_group}-{printed}.pdf"',
            "X-Printed-Count": str(printed),
            "X-Skipped-Count": str(skipped),
            # CORS/axios: чтобы фронт мог читать кастомные заголовки
            "Access-Control-Expose-Headers": "X-Printed-Count, X-Skipped-Count",
        },
    )


@router.delete("/{group_id}/items")
def clear_kiz_group_items(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    group = (
        db.query(KizGroup)
        .filter(KizGroup.id == group_id, KizGroup.user_id == current_user.id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена.")

    deleted_pool = (
        db.query(KizPoolItem)
        .filter(KizPoolItem.group_id == group_id)
        .delete(synchronize_session=False)
    )
    deleted_errors = (
        db.query(KizParserError)
        .filter(KizParserError.group_id == group_id)
        .delete(synchronize_session=False)
    )
    db.commit()
    return {"ok": True, "deleted_pool": deleted_pool, "deleted_errors": deleted_errors}


@router.delete("/{group_id}")
def delete_kiz_group(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    group = (
        db.query(KizGroup)
        .filter(KizGroup.id == group_id, KizGroup.user_id == current_user.id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена.")

    db.delete(group)
    db.commit()
    return {"ok": True}


@router.post("/product-mappings")
def upsert_product_mapping(
    payload: ProductGroupMappingUpsert,
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    article = (payload.article or "").strip()
    if not article:
        raise HTTPException(status_code=400, detail="article обязателен.")

    size = (payload.size or "").strip()

    group = (
        db.query(KizGroup)
        .filter(KizGroup.id == payload.group_id, KizGroup.user_id == current_user.id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена.")

    marketplace = (
        db.query(Marketplace)
        .filter(
            Marketplace.id == payload.marketplace_id,
            Marketplace.user_id == current_user.id,
        )
        .first()
    )
    if not marketplace:
        raise HTTPException(status_code=404, detail="Магазин не найден.")

    row = (
        db.query(KizProductMapping)
        .filter(
            KizProductMapping.user_id == current_user.id,
            KizProductMapping.marketplace_id == payload.marketplace_id,
            KizProductMapping.article == article,
            KizProductMapping.size == size,
        )
        .first()
    )
    if not row:
        row = KizProductMapping(
            user_id=current_user.id,
            marketplace_id=payload.marketplace_id,
            article=article,
            size=size,
            group_id=payload.group_id,
        )
        db.add(row)
    else:
        row.group_id = payload.group_id
    db.commit()
    return {"ok": True}


@router.get("/products/export")
def export_products_for_mapping(
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    order_size_col = func.coalesce(Order.extra_data["size"].as_string(), "")
    products = (
        db.query(
            Order.marketplace_id,
            Marketplace.name.label("marketplace_name"),
            Marketplace.type.label("marketplace_type"),
            Order.article,
            func.max(Order.product_name).label("product_name"),
            order_size_col.label("order_size"),
        )
        .join(Marketplace, Marketplace.id == Order.marketplace_id)
        .filter(Marketplace.user_id == current_user.id)
        .group_by(
            Order.marketplace_id,
            Marketplace.name,
            Marketplace.type,
            Order.article,
            order_size_col,
        )
        .order_by(Marketplace.name.asc(), Order.article.asc(), order_size_col.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(
        [
            "marketplace_id",
            "marketplace_name",
            "article",
            "color",
            "size",
            "product_name",
            "group_id",
            "group_name",
        ]
    )

    seen_export: set[tuple[int, str, str, str]] = set()
    color_markers = _normalize_color_markers(
        _get_or_create_kiz_settings(db, current_user.id).color_markers
    ) or list(DEFAULT_COLOR_MARKERS)
    for row in products:
        article_base, size, color = _article_display_parts(
            row.article,
            marketplace_type=row.marketplace_type,
            size_from_order=row.order_size,
            color_markers=color_markers,
        )
        key = (row.marketplace_id, article_base, color, size)
        if key in seen_export:
            continue
        seen_export.add(key)
        mapping = _find_product_mapping(
            db,
            user_id=current_user.id,
            marketplace_id=row.marketplace_id,
            article=article_base,
            size=size,
            color=color,
        )
        group_id = mapping[0].group_id if mapping else ""
        group_name = mapping[1] if mapping else ""
        ws.append(
            [
                row.marketplace_id,
                row.marketplace_name,
                article_base,
                color,
                size,
                row.product_name or "",
                group_id,
                group_name,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kiz-product-mapping.xlsx"'},
    )


def _parse_mapping_rows(file_name: str, content: bytes) -> list[dict[str, str]]:
    ext = Path(file_name).suffix.lower()
    if ext == ".csv":
        raw = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(raw))
        return [{k: (v or "").strip() for k, v in row.items()} for row in reader]
    if ext == ".xlsx":
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip() for c in ws[1]]
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item: dict[str, str] = {}
            for idx, key in enumerate(header):
                if not key:
                    continue
                item[key] = str(row[idx] if idx < len(row) and row[idx] is not None else "").strip()
            rows.append(item)
        wb.close()
        return rows
    raise ValueError("Поддерживаются только .xlsx или .csv.")


@router.post("/products/import")
async def import_product_mapping_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    file_name = file.filename or "mapping.xlsx"
    content = await file.read()

    try:
        rows = _parse_mapping_rows(file_name, content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    updated = 0
    created = 0
    skipped = 0

    groups_by_id = {
        g.id: g
        for g in db.query(KizGroup).filter(KizGroup.user_id == current_user.id).all()
    }
    groups_by_name = {g.name.strip().lower(): g for g in groups_by_id.values()}
    marketplaces = {
        m.id: m
        for m in db.query(Marketplace).filter(Marketplace.user_id == current_user.id).all()
    }

    for row in rows:
        article = (row.get("article") or "").strip()
        if not article:
            skipped += 1
            continue

        try:
            marketplace_id = int((row.get("marketplace_id") or "").strip())
        except Exception:
            skipped += 1
            continue
        if marketplace_id not in marketplaces:
            skipped += 1
            continue

        group = None
        group_id_raw = (row.get("group_id") or "").strip()
        if group_id_raw:
            try:
                group = groups_by_id.get(int(group_id_raw))
            except Exception:
                group = None
        if not group:
            group_name = (row.get("group_name") or "").strip().lower()
            group = groups_by_name.get(group_name)
        if not group:
            skipped += 1
            continue

        size = (row.get("size") or "").strip()
        existing = (
            db.query(KizProductMapping)
            .filter(
                KizProductMapping.user_id == current_user.id,
                KizProductMapping.marketplace_id == marketplace_id,
                KizProductMapping.article == article,
                KizProductMapping.size == size,
            )
            .first()
        )
        if existing:
            if existing.group_id != group.id:
                existing.group_id = group.id
                updated += 1
        else:
            db.add(
                KizProductMapping(
                    user_id=current_user.id,
                    marketplace_id=marketplace_id,
                    article=article,
                    size=size,
                    group_id=group.id,
                )
            )
            created += 1

    db.commit()
    return {"ok": True, "created": created, "updated": updated, "skipped": skipped}


@router.get("/reports")
def download_kiz_report(
    report_type: str = Query("free", description="free | used | errors"),
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    kind = (report_type or "").strip().lower()
    wb = Workbook()
    ws = wb.active

    if kind == "errors":
        ws.title = "errors"
        ws.append(["group_name", "source_filename", "source_page", "error_message", "created_at"])
        rows = (
            db.query(KizParserError, KizGroup.name)
            .join(KizGroup, KizGroup.id == KizParserError.group_id)
            .filter(KizParserError.user_id == current_user.id)
            .order_by(KizParserError.created_at.desc())
            .all()
        )
        for err, group_name in rows:
            ws.append(
                [
                    group_name,
                    err.source_filename or "",
                    err.source_page or "",
                    err.error_message,
                    err.created_at.isoformat(),
                ]
            )
        filename = "kiz-report-errors.xlsx"
    else:
        target_status = KizCodeStatus.FREE if kind == "free" else KizCodeStatus.USED
        ws.title = target_status.value
        ws.append(
            [
                "group_name",
                "code",
                "source_filename",
                "source_page",
                "used_order_id",
                "used_at",
            ]
        )
        rows = (
            db.query(KizPoolItem, KizGroup.name)
            .join(KizGroup, KizGroup.id == KizPoolItem.group_id)
            .filter(
                KizGroup.user_id == current_user.id,
                KizPoolItem.status == target_status,
            )
            .order_by(KizPoolItem.id.asc())
            .all()
        )
        for item, group_name in rows:
            ws.append(
                [
                    group_name,
                    item.code,
                    item.source_filename or "",
                    item.source_page or "",
                    item.used_order_id or "",
                    item.used_at.isoformat() if item.used_at else "",
                ]
            )
        filename = f"kiz-report-{target_status.value}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/products")
def list_products_for_mapping(
    search: str = Query("", description="Поиск по артикулу/названию"),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: User = CurrentAdminUser,
):
    order_size_col = func.coalesce(Order.extra_data["size"].as_string(), "")
    q = (
        db.query(
            Order.marketplace_id,
            Marketplace.name.label("marketplace_name"),
            Marketplace.type.label("marketplace_type"),
            Order.article,
            func.max(Order.product_name).label("product_name"),
            order_size_col.label("order_size"),
        )
        .join(Marketplace, Marketplace.id == Order.marketplace_id)
        .filter(Marketplace.user_id == current_user.id)
        .group_by(
            Order.marketplace_id,
            Marketplace.name,
            Marketplace.type,
            Order.article,
            order_size_col,
        )
        .order_by(Marketplace.name.asc(), Order.article.asc(), order_size_col.asc())
    )
    if search.strip():
        term = f"%{search.strip()}%"
        q = q.filter(or_(Order.article.ilike(term), Order.product_name.ilike(term)))
    rows = q.limit(limit).all()

    out = []
    seen: set[tuple[int, str, str, str]] = set()
    color_markers = _normalize_color_markers(
        _get_or_create_kiz_settings(db, current_user.id).color_markers
    ) or list(DEFAULT_COLOR_MARKERS)
    for row in rows:
        article_base, size, color = _article_display_parts(
            row.article,
            marketplace_type=row.marketplace_type,
            size_from_order=row.order_size,
            color_markers=color_markers,
        )
        key = (row.marketplace_id, article_base, color, size)
        if key in seen:
            continue
        seen.add(key)
        mapping = _find_product_mapping(
            db,
            user_id=current_user.id,
            marketplace_id=row.marketplace_id,
            article=article_base,
            size=size,
            color=color,
        )
        out.append(
            {
                "marketplace_id": row.marketplace_id,
                "marketplace_name": row.marketplace_name,
                "article": article_base,
                "color": color,
                "size": size,
                "product_name": row.product_name or "",
                "group_id": mapping[0].group_id if mapping else None,
                "group_name": mapping[1] if mapping else None,
            }
        )
    return out
