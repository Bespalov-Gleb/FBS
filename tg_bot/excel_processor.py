from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

OZON_FIRST_COLUMN = "Номер заказа"
WB_FIRST_COLUMN = "№ задания"
ROWS_PER_COLUMN = 67

# Возможные подписи столбцов в выгрузках (первое совпадение в строке заголовка побеждает).
OZON_ARTICLE_NAMES = ("Артикул",)
OZON_QTY_NAMES = ("Количество", "Кол-во", "Количество товара")
WB_SIZE_NAMES = ("Размер", "Размер товара")
WB_ARTICLE_NAMES = ("Артикул продавца", "Артикул селлера")


def _is_csv(file_path: Path) -> bool:
    return file_path.suffix.lower() == ".csv"


def _csv_matrix(file_path: Path) -> list[list[str]]:
    raw = file_path.read_text(encoding="utf-8-sig", errors="replace")
    if not raw.strip():
        return []
    sample = raw[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
    reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
    return [list(row) for row in reader]


def _header_map_from_row(header_row: list) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        normalized = _normalize_header(value)
        if normalized:
            mapping[normalized] = idx
    return mapping


def _pick_column(header_map: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for name in candidates:
        key = _normalize_header(name)
        if key and key in header_map:
            return header_map[key]
    return None


def _open_xlsx(path: Path):
    """Обычный режим чтения: при read_only=True openpyxl обрезает строки до одной ячейки в части выгрузок WB."""
    return load_workbook(path, read_only=False, data_only=True)


@dataclass(slots=True)
class SourceFile:
    path: Path
    source_type: str  # "ozon" | "wb"


def process_files(input_paths: list[Path], output_dir: Path) -> tuple[Path, dict[str, int]]:
    classified = [SourceFile(path=p, source_type=detect_source_type(p)) for p in input_paths]

    ozon_articles: list[str] = []
    wb_articles: list[str] = []

    for source in classified:
        if source.source_type == "ozon":
            ozon_articles.extend(parse_ozon_articles(source.path))
        elif source.source_type == "wb":
            wb_articles.extend(parse_wb_articles(source.path))

    workbook = Workbook()
    ozon_ws = workbook.active
    ozon_ws.title = "ozon"
    wb_ws = workbook.create_sheet("wb")

    _fill_sheet(ozon_ws, "ozon", ozon_articles)
    _fill_sheet(wb_ws, "wb", wb_articles)

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    output_path = output_dir / filename
    workbook.save(output_path)

    return output_path, {"ozon": len(ozon_articles), "wb": len(wb_articles)}


def detect_source_type(file_path: Path) -> str:
    if _is_csv(file_path):
        matrix = _csv_matrix(file_path)
        if not matrix or not matrix[0]:
            raise ValueError(f"Файл {file_path.name}: пустой или некорректный CSV.")
        first_cell = matrix[0][0] if matrix[0] else None
        first_header = _normalize_header(first_cell)
    else:
        wb = _open_xlsx(file_path)
        ws = wb.active
        first_header = _normalize_header(ws.cell(row=1, column=1).value)
        wb.close()

    if first_header == _normalize_header(OZON_FIRST_COLUMN):
        return "ozon"
    if first_header == _normalize_header(WB_FIRST_COLUMN):
        return "wb"

    raise ValueError(
        f"Файл {file_path.name} не распознан. Первый столбец должен быть "
        f"'{OZON_FIRST_COLUMN}' или '{WB_FIRST_COLUMN}'."
    )


def parse_ozon_articles(file_path: Path) -> list[str]:
    if _is_csv(file_path):
        matrix = _csv_matrix(file_path)
        if len(matrix) < 2:
            raise ValueError(
                f"Файл {file_path.name}: в CSV нет строк данных под заголовком."
            )
        header_map = _header_map_from_row(matrix[0])
        data_iter = matrix[1:]
    else:
        wb = _open_xlsx(file_path)
        ws = wb.active
        header_map = _header_map(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        data_iter = ws.iter_rows(min_row=2, values_only=True)

    article_idx = _pick_column(header_map, OZON_ARTICLE_NAMES)
    qty_idx = _pick_column(header_map, OZON_QTY_NAMES)
    if article_idx is None or qty_idx is None:
        if not _is_csv(file_path):
            wb.close()
        raise ValueError(
            f"Файл {file_path.name}: не найдены столбцы Ozon. Нужны артикул "
            f"({', '.join(OZON_ARTICLE_NAMES)}) и количество ({', '.join(OZON_QTY_NAMES)})."
        )

    rows: list[tuple[str, int]] = []
    for row in data_iter:
        if row is None:
            continue
        row_list = list(row) if not isinstance(row, list) else row
        article_raw = row_list[article_idx] if article_idx < len(row_list) else None
        qty_raw = row_list[qty_idx] if qty_idx < len(row_list) else None

        article = str(article_raw).strip() if article_raw is not None else ""
        if not article:
            continue

        qty = _parse_qty(qty_raw)
        rows.append((article, qty))

    if not _is_csv(file_path):
        wb.close()

    current_count = len(rows)
    qty_sum = sum(qty for _, qty in rows)

    if qty_sum != current_count and qty_sum > current_count:
        expanded: list[str] = []
        for article, qty in rows:
            expanded.extend([article] * qty)
        return expanded

    return [article for article, _ in rows]


def parse_wb_articles(file_path: Path) -> list[str]:
    if _is_csv(file_path):
        matrix = _csv_matrix(file_path)
        if len(matrix) < 2:
            raise ValueError(
                f"Файл {file_path.name}: в CSV нет строк данных под заголовком."
            )
        header_map = _header_map_from_row(matrix[0])
        data_iter = matrix[1:]
    else:
        wb = _open_xlsx(file_path)
        ws = wb.active
        header_map = _header_map(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        data_iter = ws.iter_rows(min_row=2, values_only=True)

    size_idx = _pick_column(header_map, WB_SIZE_NAMES)
    article_idx = _pick_column(header_map, WB_ARTICLE_NAMES)
    if size_idx is None or article_idx is None:
        if not _is_csv(file_path):
            wb.close()
        raise ValueError(
            f"Файл {file_path.name}: не найдены столбцы WB. Нужны размер "
            f"({', '.join(WB_SIZE_NAMES)}) и артикул продавца ({', '.join(WB_ARTICLE_NAMES)})."
        )

    result: list[str] = []
    for row in data_iter:
        if row is None:
            continue
        row_list = list(row) if not isinstance(row, list) else row
        article_raw = row_list[article_idx] if article_idx < len(row_list) else None
        size_raw = row_list[size_idx] if size_idx < len(row_list) else None

        article = str(article_raw).strip() if article_raw is not None else ""
        size = str(size_raw).strip() if size_raw is not None else ""

        if not article and not size:
            continue
        result.append(f"{article}_{size}".strip("_"))

    if not _is_csv(file_path):
        wb.close()
    return result


def _fill_sheet(ws, sheet_label: str, articles: list[str]) -> None:
    sorted_articles = sorted(articles, key=lambda x: x.lower())
    white_articles = [a for a in sorted_articles if "white" in a.lower()]
    main_articles = [a for a in sorted_articles if "white" not in a.lower()]

    ws["A1"] = sheet_label
    ws["B1"] = len(sorted_articles)

    used_columns = _write_blocks(ws, start_column=1, values=main_articles)
    white_start = used_columns + 1
    ws.cell(row=1, column=white_start, value=f"{sheet_label}_white")
    _write_blocks(ws, start_column=white_start, values=white_articles)

    _autofit_columns(ws)
    _apply_print_setup(ws)


def _write_blocks(ws, start_column: int, values: Iterable[str]) -> int:
    values = list(values)
    if not values:
        return start_column

    for idx, value in enumerate(values):
        block = idx // ROWS_PER_COLUMN
        row_in_block = (idx % ROWS_PER_COLUMN) + 2
        column = start_column + block
        ws.cell(row=row_in_block, column=column, value=value)

    return start_column + ((len(values) - 1) // ROWS_PER_COLUMN)


def _autofit_columns(ws, min_width: float = 8.0, max_width: float = 72.0) -> None:
    """Подбор ширины столбцов по длине текста (openpyxl не умеет настоящий Excel AutoFit)."""
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        longest = 0
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                longest = max(longest, len(str(val)))
        # Небольшой запас под отступы и шрифт по умолчанию
        width = min(max(longest + 2.5, min_width), max_width)
        ws.column_dimensions[letter].width = width


def _apply_print_setup(ws) -> None:
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def _header_map(header_rows) -> dict[str, int]:
    values = next(header_rows, ())
    mapping: dict[str, int] = {}
    for idx, value in enumerate(values):
        normalized = _normalize_header(value)
        if normalized:
            mapping[normalized] = idx
    return mapping


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_qty(value) -> int:
    if value is None or str(value).strip() == "":
        return 1
    try:
        parsed = int(float(str(value).replace(",", ".")))
    except ValueError:
        return 1
    return max(parsed, 1)
